from openpyxl.styles import Border, Side, PatternFill, Font
import pandas as pd
from openpyxl.utils import get_column_letter


def cemix_Synthese_to_excel(df, export):
    
    excel_writer = pd.ExcelWriter(export, engine='openpyxl')

    df.to_excel(excel_writer, sheet_name='Sheet1', index=False)
    cols = list(df.columns)

    workbook = excel_writer.book
    worksheet = excel_writer.sheets['Sheet1']

    for idx, column in enumerate(df.columns, 1):
        col_letter = get_column_letter(idx)
        worksheet.column_dimensions[col_letter].width = 17

    border = Border(left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin'))

    colors = ["FFFFFF" for i in df.columns] 
    for i in cols:
        if cols.index(i)%2:
            colors[cols.index(i)] = "CCCCCC"

    fills = [PatternFill(start_color=color, end_color=color, fill_type='solid') for color in colors]

    font = Font(bold=True, color='000000')

    for idx, column in enumerate(df.columns, 0):
        for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1, min_col=idx, max_col=idx):
            for cell in row:
                cell.border = border
                cell.fill = fills[idx - 1]
                cell.font = font

    # Save the Excel file
    excel_writer.save()
    print("The File is successfully exported in: ", export)
