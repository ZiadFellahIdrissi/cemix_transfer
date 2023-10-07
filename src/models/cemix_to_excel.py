import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def cemix_to_excel(df_header, df_, header_name, filename, ishift, size, num_rows):

    # Create a new workbook and select the active sheet
    wb = openpyxl.Workbook()
    sheet = wb.active
    
    # Define the column pairs and their corresponding values
    column_pairs = [("A", "B"), ("C", "D"), ("E", "F"), ("G", "H"), ("I", "J")]

    # Set the width of each cell to 16 pixels
    for column in column_pairs:
        sheet.column_dimensions[column[0]].width = size
        sheet.column_dimensions[column[1]].width = size

    # Create the header
    header_row = 1
    header_title = header_name
    header_font = Font(size = 17, bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")
    header_range = f"A{header_row}:{get_column_letter(2 * len(column_pairs))}{header_row + 1}"
    
    header_cell = sheet[header_range.split(":")[0]]
    header_cell.value = header_title
    header_cell.font = header_font
    header_cell.alignment = header_alignment
    header_cell.fill = header_fill
    sheet.merge_cells(header_range)

    
    start_row = 4
    indx = 0
    for row_index in range(start_row, num_rows + start_row):
        for col1, col2 in column_pairs:

            cell_range = f'{col1}{row_index}:{col2}{row_index}'
            sheet.merge_cells(cell_range)
            merged_cell = sheet[cell_range.split(":")[0]]
            text = df_header["Variable"][indx] + " : " + df_header["Value"][indx]
            merged_cell.value = text

            # Set the background color based on the column pair index
            if column_pairs.index((col1, col2)) % 2 == 0:
                # Use the first color (#cccccc) for even index column pairs
                merged_cell.fill = PatternFill(start_color="cccccc", end_color="cccccc", fill_type="solid")
            else:
                # Use the second color (#efefef) for odd index column pairs
                merged_cell.fill = PatternFill(start_color="efefef", end_color="efefef", fill_type="solid")


            indx = indx + 1


    # Insert the DataFrame into the Excel sheet
    if df_ is not None and ishift == True:
        for r_idx, row in enumerate(df_.iterrows(), row_index + 1): 
            for c_idx, value in enumerate(row[1], 1):
                
                if c_idx == 5:  # Check if it's the 5th column
                    sheet.merge_cells(start_row=r_idx, start_column=c_idx, end_row=r_idx, end_column=c_idx + 1)
                    sheet.cell(row=r_idx, column=c_idx, value=value)

                elif c_idx < 5:
                    sheet.cell(row=r_idx, column=c_idx, value=value)

                else:
                    sheet.cell(row=r_idx, column=c_idx+1, value=value)

    elif ishift == False:
        for r_idx, row in enumerate(df_.iterrows(), row_index + 1): 
            for c_idx, value in enumerate(row[1], 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)


    
    # Add an outside border to all cells
    border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border
    
    # Save the workbook to the specified filename
    wb.save(filename)
    print(f'Excel file "{filename}" created successfully.')
